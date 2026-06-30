#!/usr/bin/env python3
import streamlit as st
import tempfile, os, io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from detect_duplicates import (
    extract_pages, parse_questions, find_duplicates,
    group_duplicates, make_desc, fmt_qs, fmt_pages, is_figure_based
)

st.set_page_config(
    page_title="Duplicate Question Detector",
    page_icon="🔍",
    layout="wide"
)

# Keep results in session state so download clicks don't wipe them
if "groups" not in st.session_state:
    st.session_state.groups    = None
    st.session_state.questions = None
    st.session_state.txt_bytes = None
    st.session_state.xlsx_bytes = None

st.title("Duplicate Question Detector")
st.caption("Upload a question bank PDF to find duplicate and near-duplicate questions across any subject.")

uploaded = st.file_uploader("Upload PDF", type=["pdf"])

if uploaded:
    if st.button("Detect Duplicates", type="primary"):

        # Clear previous results when a new run starts
        st.session_state.groups     = None
        st.session_state.questions  = None
        st.session_state.txt_bytes  = None
        st.session_state.xlsx_bytes = None

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded.read())
            tmp_path = tmp.name

        try:
            with st.status("Processing...", expanded=True) as status:

                st.write("Extracting text from PDF...")
                pages = extract_pages(tmp_path)
                st.write(f"  {len(pages)} pages extracted")

                st.write("Parsing questions...")
                questions = parse_questions(pages)
                st.write(f"  {len(questions)} questions found")

                if not questions:
                    status.update(label="Failed", state="error")
                    st.error("No questions found. Make sure the PDF has questions numbered Q1, Q2, ...")
                    st.stop()

                st.write("Detecting duplicates (this may take 30-60 seconds for large PDFs)...")
                pairs = find_duplicates(questions)
                exact = sum(1 for p in pairs if p[3] == 'Exact')
                near  = sum(1 for p in pairs if p[3] == 'Near-Exact')
                st.write(f"  {exact} exact + {near} near-duplicate pairs found")

                st.write("Grouping into clusters...")
                groups = group_duplicates(pairs)

                status.update(
                    label=f"Done — {len(groups)} duplicate groups found",
                    state="complete"
                )

            # ── Build text report ──────────────────────────────────────────────
            lines = ["Direct Text and Data Duplicates", "=" * 60, ""]
            for i, g in enumerate(groups, 1):
                qs         = g['questions']
                pages_list = [questions[q]['page'] for q in qs]
                desc       = make_desc(questions[qs[0]]['text'])
                note       = " [Figure/graph-based: verify manually]" \
                             if any(is_figure_based(questions[q]['text']) for q in qs) else ""
                lines.append(f"{i}. {fmt_qs(qs)} ({fmt_pages(pages_list)}): {desc}.{note}")
            lines.append(f"\nTotal: {len(groups)} duplicate groups found")
            txt_bytes = "\n".join(lines).encode("utf-8")

            # ── Build Excel report ─────────────────────────────────────────────
            xlsx_buf  = io.BytesIO()
            wb        = Workbook()
            hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            hdr_font  = Font(color="FFFFFF", bold=True)

            ws1 = wb.active
            ws1.title = "Summary"
            for c, h in enumerate(["#", "Questions", "Pages", "Similarity %", "Type", "Description"], 1):
                cell = ws1.cell(1, c, h)
                cell.fill, cell.font = hdr_fill, hdr_font
                cell.alignment = Alignment(horizontal="center")

            for i, g in enumerate(groups, 1):
                qs         = g['questions']
                pages_list = [questions[q]['page'] for q in qs]
                fill       = PatternFill("solid", fgColor=("EBF2FF" if i % 2 else "FFFFFF"))
                row        = [i, fmt_qs(qs), fmt_pages(pages_list),
                              round(g['score']), g['type'],
                              make_desc(questions[qs[0]]['text'])]
                for c, v in enumerate(row, 1):
                    cell = ws1.cell(i + 1, c, v)
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True)

            for col, w in zip("ABCDEF", [6, 32, 22, 14, 14, 65]):
                ws1.column_dimensions[col].width = w

            ws2 = wb.create_sheet("Question Details")
            for c, h in enumerate(["Group #", "Question", "Page", "Full Text (first 400 chars)"], 1):
                cell = ws2.cell(1, c, h)
                cell.fill, cell.font = hdr_fill, hdr_font

            r = 2
            for i, g in enumerate(groups, 1):
                fill = PatternFill("solid", fgColor=("EBF2FF" if i % 2 else "FFFFFF"))
                for qn in g['questions']:
                    vals = [i, f"Q{qn}", questions[qn]['page'], questions[qn]['text'][:400]]
                    for c, v in enumerate(vals, 1):
                        cell = ws2.cell(r, c, v)
                        cell.fill = fill
                        cell.alignment = Alignment(wrap_text=True)
                    r += 1

            for col, w in zip("ABCD", [8, 10, 8, 85]):
                ws2.column_dimensions[col].width = w

            wb.save(xlsx_buf)

            # Save everything to session state
            st.session_state.groups     = groups
            st.session_state.questions  = questions
            st.session_state.txt_bytes  = txt_bytes
            st.session_state.xlsx_bytes = xlsx_buf.getvalue()

        finally:
            os.unlink(tmp_path)

# ── Show results from session state (persists across download clicks) ──────────
if st.session_state.groups is not None:
    groups    = st.session_state.groups
    questions = st.session_state.questions

    st.subheader(f"Results: {len(groups)} Duplicate Groups")

    rows = []
    for i, g in enumerate(groups, 1):
        qs         = g['questions']
        pages_list = [questions[q]['page'] for q in qs]
        rows.append({
            "#":            i,
            "Questions":    fmt_qs(qs),
            "Pages":        fmt_pages(pages_list),
            "Similarity %": round(g['score']),
            "Type":         g['type'],
            "Description":  make_desc(questions[qs[0]]['text']),
        })

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="Download Text Report (.txt)",
            data=st.session_state.txt_bytes,
            file_name="duplicate_report.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with col2:
        st.download_button(
            label="Download Excel Report (.xlsx)",
            data=st.session_state.xlsx_bytes,
            file_name="duplicate_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
