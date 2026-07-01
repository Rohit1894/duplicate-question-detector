#!/usr/bin/env python3
"""
Duplicate Question Detector
Finds exact and near-duplicate questions in a JEE question bank PDF.
Outputs: duplicate_report.txt  +  duplicate_report.xlsx
"""

import pdfplumber
import re
import hashlib
import os
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── Configuration ────────────────────────────────────────────────────────────────
PDF_PATH  = r"E:\ai automation\Duplicate questions\input\Motion in a Straight Line_JEE Question Bank 2027 (5).pdf"
OUT_DIR   = r"E:\ai automation\Duplicate questions\output"
TXT_OUT   = os.path.join(OUT_DIR, "duplicate_report.txt")
XLSX_OUT  = os.path.join(OUT_DIR, "duplicate_report_v2.xlsx")
THRESHOLD = 85   # Similarity % for near-duplicate detection (0–100)


# ── Step 1: PDF Extraction ───────────────────────────────────────────────────────
def extract_pages(pdf_path):
    """Extract text per page. Splits each page at midpoint to handle two-column layout."""
    pages = {}
    with pdfplumber.open(pdf_path) as pdf:
        for num, page in enumerate(pdf.pages, 1):
            w, h = page.width, page.height
            left  = page.within_bbox((0,   0, w/2, h)).extract_text(x_tolerance=3, y_tolerance=3) or ""
            right = page.within_bbox((w/2, 0, w,   h)).extract_text(x_tolerance=3, y_tolerance=3) or ""
            pages[num] = left + "\n" + right
    return pages


# ── Step 2: Question Parsing ─────────────────────────────────────────────────────
def parse_questions(pages):
    """Return {q_num: {'text': ..., 'page': ...}} from all pages."""
    # Embed page markers so we can trace which page each question starts on
    tagged = "\n".join(f"<<<P:{n}>>>\n{t}" for n, t in sorted(pages.items()))

    # Match Q1. or Q1) or Q1 (word boundary at start, digit-terminating char after number)
    q_re = re.compile(r'\bQ(\d{1,4})[\.)\s]', re.IGNORECASE)
    hits = [(m.start(), int(m.group(1))) for m in q_re.finditer(tagged)]

    questions = {}
    for i, (pos, qnum) in enumerate(hits):
        end     = hits[i + 1][0] if i + 1 < len(hits) else len(tagged)
        segment = tagged[pos:end]

        # Last page marker before this position → the page this question starts on
        prior      = tagged[:pos]
        page_hits  = re.findall(r'<<<P:(\d+)>>>', prior)
        page       = int(page_hits[-1]) if page_hits else 1

        # Strip page markers and leading Q-prefix from text
        text = re.sub(r'<<<P:\d+>>>', ' ', segment)
        text = re.sub(r'^\bQ\d+[\.)\s]+', '', text, count=1)
        text = re.sub(r'\s+', ' ', text).strip()

        # Skip answer-key entries (very short or all uppercase single chars)
        if len(text) < 20:
            continue
        if re.fullmatch(r'[A-D\d,\.\s]+', text):
            continue

        # First occurrence wins (don't overwrite if Q-number appears in answer key again)
        if qnum > 0 and qnum not in questions:
            questions[qnum] = {'text': text, 'page': page}

    return questions


# ── Step 3: Text Normalization ───────────────────────────────────────────────────
def normalize(text):
    """Strip MCQ option labels, blanks, watermarks, punctuation; lowercase."""
    text = re.sub(r'\([A-Da-d]\)\s*', ' ', text)                          # (A), (B)
    text = re.sub(r'\b[A-Da-d]\.\s+', ' ', text)                          # A. B.
    text = re.sub(r'_{2,}', 'BLANK', text)                                # ___
    text = re.sub(r'master\s+nce{1,2}r?\s+jee', ' ', text, flags=re.IGNORECASE) # watermark
    text = re.sub(r'pw\s+books?\s+app', ' ', text, flags=re.IGNORECASE)         # watermark
    text = re.sub(r'[^\w\s]', ' ', text)                                  # punctuation
    return re.sub(r'\s+', ' ', text.lower()).strip()


# ── Step 4: Duplicate Detection ──────────────────────────────────────────────────
def find_duplicates(questions, threshold=THRESHOLD):
    """Return list of (q_a, q_b, score, type) for all duplicate pairs."""
    norm  = {qn: normalize(d['text']) for qn, d in questions.items()}
    pairs = []

    # Layer 1 — exact hash match
    buckets = {}
    for qn, t in norm.items():
        buckets.setdefault(hashlib.md5(t.encode()).hexdigest(), []).append(qn)

    exact_set = set()
    for nums in buckets.values():
        if len(nums) > 1:
            for ii in range(len(nums)):
                for jj in range(ii + 1, len(nums)):
                    a, b = nums[ii], nums[jj]
                    pairs.append((a, b, 100, 'Exact'))
                    exact_set.add((min(a, b), max(a, b)))

    # Layer 2 — fuzzy match on remaining pairs
    qlist = sorted(questions)
    n     = len(qlist)
    print(f"         Running fuzzy check on {n*(n-1)//2:,} pairs ...", end="", flush=True)

    checked = 0
    for i in range(n):
        for j in range(i + 1, n):
            qi, qj = qlist[i], qlist[j]
            if (min(qi, qj), max(qi, qj)) in exact_set:
                continue
            ti, tj = norm[qi], norm[qj]
            if not ti or not tj:
                continue
            li, lj = len(ti), len(tj)
            # Skip pairs whose length ratio is too different — can't be 85%+ similar
            if min(li, lj) / max(li, lj) < 0.45:
                continue
            # Skip very short normalized texts — generic headers like "match list i with list ii"
            # will falsely match; for short texts only exact hash match is reliable
            if min(li, lj) < 40:
                continue
            sort_score = fuzz.token_sort_ratio(ti, tj)
            set_score  = fuzz.token_set_ratio(ti, tj)
            score      = max(sort_score, set_score)
            # "Match List" questions share a long generic stem that inflates scores;
            # require near-identical text (99%) before flagging them as duplicates
            is_ml_pair = (MATCH_LIST_RE.search(questions[qi]['text']) is not None and
                          MATCH_LIST_RE.search(questions[qj]['text']) is not None)
            if is_ml_pair:
                if sort_score >= 99 or set_score >= 99:
                    pairs.append((qi, qj, score, 'Near-Exact'))
                continue
            # Three acceptance conditions for normal questions:
            # 1. High token_sort (normal case)
            # 2. Very high token_set (clean near-duplicate caught by set overlap)
            # 3. Both metrics moderately high — targets garbled PDF column text where
            #    word order is scrambled but the right tokens are all present
            both_moderate = sort_score >= (threshold - 13) and set_score >= (threshold - 1)
            if sort_score >= threshold or set_score >= (threshold + 5) or both_moderate:
                # Veto: same structure but different physics values = different question
                # e.g. "car at 20 km/h" vs "car at 40 km/h" → skip even if 90%+ similar
                # Uses subset check (not equality) to tolerate PDF extraction artifacts
                nums_i = extract_stem_numbers(questions[qi]['text'])
                nums_j = extract_stem_numbers(questions[qj]['text'])
                if (nums_i and nums_j and
                        not nums_i.issubset(nums_j) and
                        not nums_j.issubset(nums_i)):
                    continue
                pairs.append((qi, qj, score, 'Near-Exact'))
            checked += 1

    print(f" done ({checked:,} checked)")
    return pairs


# ── Step 5: Group Pairs into Clusters ────────────────────────────────────────────
def group_duplicates(pairs):
    """Union-find: merge overlapping pairs into clusters (handles triples, etc.)."""
    parent = {}
    info   = {}   # (min_q, max_q) -> (score, type)

    def find(x):
        parent.setdefault(x, x)
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]

    for a, b, score, dtype in pairs:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb
        info[(min(a, b), max(a, b))] = (score, dtype)

    clusters = {}
    for node in parent:
        clusters.setdefault(find(node), set()).add(node)

    groups = []
    for members in clusters.values():
        if len(members) < 2:
            continue
        members = sorted(members)
        pair_scores = [info[(min(a, b), max(a, b))][0]
                       for ii, a in enumerate(members)
                       for b in members[ii + 1:]
                       if (min(a, b), max(a, b)) in info]
        pair_types  = [info[(min(a, b), max(a, b))][1]
                       for ii, a in enumerate(members)
                       for b in members[ii + 1:]
                       if (min(a, b), max(a, b)) in info]
        groups.append({
            'questions': members,
            'score': max(pair_scores) if pair_scores else 0,
            'type':  'Exact' if all(t == 'Exact' for t in pair_types) else 'Near-Exact',
        })

    return sorted(groups, key=lambda g: g['questions'][0])


# ── Helpers ───────────────────────────────────────────────────────────────────────
FIGURE_RE = re.compile(
    r'\b(figure|graph|diagram|shown in|see fig|from the following|as shown)\b',
    re.IGNORECASE
)

MATCH_LIST_RE = re.compile(r'\bmatch\s+(list|column)', re.IGNORECASE)


def extract_stem_numbers(text):
    """Extract large numeric values (>=10) from question stem only (before MCQ options).
    Only large numbers are meaningful physics values; small numbers (1,2,3) appear
    incidentally everywhere and cause false vetoes."""
    stem = re.split(r'\s*\([A-Da-d]\)\s*', text)[0]
    return {n for n in re.findall(r'\d+(?:\.\d+)?', stem) if float(n) >= 10}


def make_desc(text, max_chars=130):
    """Short human-readable description from question text (strips option section)."""
    clean = re.sub(r'Master\s+NCE{1,2}R?\s+JEE', '', text)
    clean = re.sub(r'PW\s+Books?\s+App', '', clean, flags=re.IGNORECASE)
    clean = re.sub(r'\s*\([A-Da-d]\).*', '', clean, flags=re.DOTALL).strip()
    if not clean:
        clean = text
    clean = re.sub(r'\s+', ' ', clean).strip()
    if len(clean) > max_chars:
        clean = clean[:max_chars].rsplit(' ', 1)[0] + '...'
    return clean


def is_figure_based(text):
    return bool(FIGURE_RE.search(text))


def fmt_qs(nums):
    qs = [f"Q{n}" for n in nums]
    if len(qs) == 2:
        return f"{qs[0]} and {qs[1]}"
    return ", ".join(qs[:-1]) + f", and {qs[-1]}"


def fmt_pages(page_list):
    ps = sorted(set(page_list))
    if len(ps) == 1:
        return f"Page {ps[0]}"
    if len(ps) == 2:
        return f"Pages {ps[0]} & {ps[1]}"
    return "Pages " + ", ".join(str(p) for p in ps[:-1]) + f" & {ps[-1]}"


# ── Output: Text Report ───────────────────────────────────────────────────────────
def write_text(groups, questions, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write("Direct Text and Data Duplicates\n")
        f.write("=" * 60 + "\n\n")
        for i, g in enumerate(groups, 1):
            qs    = g['questions']
            pages = [questions[q]['page'] for q in qs]
            desc  = make_desc(questions[qs[0]]['text'])
            note  = " [Figure/graph-based: verify manually]" \
                    if any(is_figure_based(questions[q]['text']) for q in qs) else ""
            f.write(f"{i}. {fmt_qs(qs)} ({fmt_pages(pages)}): {desc}.{note}\n")
        f.write(f"\nTotal: {len(groups)} duplicate groups found\n")
    print(f"  Text  -> {path}")


# ── Output: Excel Report ──────────────────────────────────────────────────────────
def write_excel(groups, questions, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    for c, h in enumerate(["#", "Questions", "Pages", "Similarity %", "Type", "Description"], 1):
        cell = ws1.cell(1, c, h)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(horizontal='center')

    for i, g in enumerate(groups, 1):
        qs    = g['questions']
        pages = [questions[q]['page'] for q in qs]
        fill  = PatternFill("solid", fgColor=("EBF2FF" if i % 2 else "FFFFFF"))
        row   = [i, fmt_qs(qs), fmt_pages(pages), round(g['score']), g['type'],
                 make_desc(questions[qs[0]]['text'])]
        for c, v in enumerate(row, 1):
            cell = ws1.cell(i + 1, c, v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    for col, w in zip("ABCDEF", [6, 32, 22, 14, 14, 65]):
        ws1.column_dimensions[col].width = w

    # ---- Sheet 2: Full Question Text ----
    ws2 = wb.create_sheet("Question Details")
    for c, h in enumerate(["Group #", "Question", "Page", "Full Text (first 400 chars)"], 1):
        cell = ws2.cell(1, c, h)
        cell.fill, cell.font = hdr_fill, hdr_font

    r = 2
    for i, g in enumerate(groups, 1):
        fill = PatternFill("solid", fgColor=("EBF2FF" if i % 2 else "FFFFFF"))
        for qn in g['questions']:
            vals = [i, f"Q{qn}", questions[qn]['page'], questions[qn]['text'][:400]]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r, c, v)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)
            r += 1

    for col, w in zip("ABCD", [8, 10, 8, 85]):
        ws2.column_dimensions[col].width = w

    wb.save(path)
    print(f"  Excel -> {path}")


# ── Main ──────────────────────────────────────────────────────────────────────────
def main():
    print("\n== Duplicate Question Detector ==========================")

    print("Step 1/5  Extracting text from PDF ...")
    pages = extract_pages(PDF_PATH)
    print(f"         {len(pages)} pages extracted")

    print("Step 2/5  Parsing questions ...")
    questions = parse_questions(pages)
    print(f"         {len(questions)} questions found (expected ~844)")

    if not questions:
        print("\nERROR: No questions parsed. Check PDF path and layout.")
        return

    # Quick sanity check
    sample_keys = sorted(questions)[:5]
    print(f"         Sample Q-numbers: {sample_keys}")

    print("Step 3/5  Detecting duplicates ...")
    pairs = find_duplicates(questions)
    exact_count = sum(1 for p in pairs if p[3] == 'Exact')
    near_count  = sum(1 for p in pairs if p[3] == 'Near-Exact')
    print(f"         {exact_count} exact pairs + {near_count} near-exact pairs = {len(pairs)} total")

    print("Step 4/5  Grouping overlapping pairs into clusters ...")
    groups = group_duplicates(pairs)
    print(f"         {len(groups)} duplicate groups")

    print("Step 5/5  Writing reports ...")
    write_text (groups, questions, TXT_OUT)
    write_excel(groups, questions, XLSX_OUT)

    print(f"\n{'='*55}")
    print(f"  {len(groups)} duplicate groups written to output/")
    print(f"{'='*55}")

    if groups:
        print("\nPreview - first 8 groups:")
        for g in groups[:8]:
            qs    = g['questions']
            pages = [questions[q]['page'] for q in qs]
            score = g['score']
            desc  = make_desc(questions[qs[0]]['text'])[:75]
            print(f"  [{round(score):3d}%] {fmt_qs(qs)} ({fmt_pages(pages)}): {desc}")


if __name__ == "__main__":
    main()
